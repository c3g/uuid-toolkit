"""
File parsing functions for uploaded files that should contain identifiers for validation or data rows to create identifiers for.

This module is responsible for taking in file bytes provided by the pipeline and then converting them into a list of python dictionaries 
that the rest of the pipeline can process,

This module only supports the following file formats:
- CSV
- JSON
- XLSX

The parser doesn't validate or clean up any of the data that the user provided.
It doesn't handle any validation or generation of identifiers.
It only takes the file and converts it into a standardized list of dictionaries.
More detailed normalization and cleanup are handled by the normalizer.py and strategy specific validation.


"""

import json
import csv
import io

from openpyxl import load_workbook


def parse_file(
        file_bytes: bytes, 
        file_type: str,
        id_name:str | None,
        sheet_name:str |None,
        ):
    """
    Parses an uploaded file into a list of dictionaries.

    Depending on the file type the function will adjust the way the file is parsed.
    Supported file types:
    - CSV
    - JSON
    - XLSX

    Each returned dictionary represetns one row from the file with the key being the column name and the values are the cell values.

    """
    if file_type == 'json':
        return json.loads(file_bytes.decode('utf-8-sig'))
    elif file_type == 'csv':
        return list(csv.DictReader(io.StringIO(file_bytes.decode('utf-8-sig'))))
    elif file_type == 'xlsx':
        return parse_xlsx(
            file_bytes,
            id_name=id_name,
            sheet_name=sheet_name,
            )
    else:
        raise ValueError(f"Unsupported file type: {file_type}")
    
def parse_xlsx(
        file_bytes:bytes,
        id_name:str | None,
        sheet_name:str |None,
    ) -> list[dict]:
    """
    Parse XLSX files bytes into a list of dictionaries

    The XLSX parser reads the uploaded workboo, selects the requested worksheet or defaults to the active worksheet.
    Extracts headers from the first row, and converts the remaining non-empty rows into dictionaries.
    """
    
    workbook = load_workbook(
        filename=io.BytesIO(file_bytes),
        data_only=True,
        read_only=True,
    )
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' was not found."
                f"Available sheets: '{workbook.sheetnames}'"
            )
        worksheet = workbook[sheet_name]

    else:
        worksheet = workbook.active
    

    rows = list(worksheet.iter_rows(values_only=True))

    if not rows:
        raise ValueError("Excel file is empty.")
    
    header_index = detect_header_row(rows,id_name = id_name)
    
    headers = rows[header_index]


    clean_headers = [
        str(header).strip() if header is not None else None
        for header in headers
    ]

    if all(header is None or header == "" for header in clean_headers):
        raise ValueError("Excel file must contain at least one non-empty header.")

    records = []

    for row in rows[header_index+1:]:
        if row is None or all(value is None for value in row):
            continue

        record = {}

        for header, value in zip(clean_headers, row):
            if header is None or header == "":
                continue

            record[header] = normalize_excel_cell(value)

        if any(value not in ("", None) for value in record.values()):
            records.append(record)

    return records

def normalize_excel_cell(value):
    if value == None:
        return ""
    return str(value)

def detect_header_row(
        rows: list[tuple],
        id_name: str | None,
) -> int:
    target_id_name = id_name.strip().lower() if id_name else None

    for index, row in enumerate(rows):
        normalized_values = [
            str(value).strip().lower()
            for value in row
            if value is not None and str(value).strip() != ""
        ]

        if not normalized_values:
            continue

        if target_id_name and target_id_name in normalized_values:
            return index

        # Helpful fallback for Genome Centre sample manifests
        if (
            "sample name" in normalized_values
            and "sample type" in normalized_values
            and "sample kind" in normalized_values
        ):
            return index

        # Generic fallback for simpler spreadsheets
        if any(
            value in {"id", "identifier", "uuid", "uid"}
            for value in normalized_values
        ):
            return index
    #Default state of using the first row as the headers
    return 0




#just testing parser logic here, not part of main codebase
if __name__ == "__main__":
    #Testing Parser logic
    # Example JSON file path (local test only)
    test_file_path = "data/sample1.json"

    # Read file as bytes (simulating upload/API behavior)
    with open(test_file_path, "rb") as f:
        file_bytes = f.read()

    #print raw bytes for debugging
    print("Raw file bytes:")
    print(file_bytes)
    # Call parser
    parsed_output = parse_file(file_bytes, file_type="json")

    # Print results clearly
    print("Parsed output:")
    print(parsed_output)
    print("Type of parsed output:", type(parsed_output))
